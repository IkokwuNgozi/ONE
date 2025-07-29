from flask import Flask, render_template, request, redirect
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

app = Flask(__name__)
logbook_filename = "new_logbook.xlsx"

# Load or create Excel workbook
if not os.path.exists(logbook_filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Logs"
    ws.append(["Date", "Time", "Name", "Activity", "Remarks"])
    wb.save(logbook_filename)

# Route to display the log form
@app.route("/", methods=["GET", "POST"])
def log_entry():
    if request.method == "POST":
        name = request.form.get("name")
        activity = request.form.get("activity")
        remarks = request.form.get("remarks")
        now = datetime.now()
        file_path = 'new_logbook.xlsx'

        # Load workbook and append data
        wb = load_workbook(logbook_filename)
        ws = wb.active
        ws.append([
            now.strftime("%Y-%m-%d"),
            now.strftime("%H:%M:%S"),
            name, activity, remarks
        ])
        wb.save(logbook_filename)

        return redirect("/success")

    return render_template("design.html")

@app.route("/success")
def success():
    return "Log entry added successfully! <a href='/'>Add another</a>"

if __name__ == "__main__":
    app.run(debug=True)

import webbrowser
webbrowser.open("http://127.0.0.1:5000/")

